#!/usr/bin/env python3
"""
Parse a supplier-quotation xlsx export (the kind that lists, per product, the
lowest/highest 1-pcs price suppliers bid) and filter it down to one Shopify
store.

Usage:
    python3 parse_quotation_xlsx.py <path/to/quotations.xlsx> --store <myshopify-subdomain> [--out parsed.json]

The xlsx is expected to have a header row with (at least) these columns —
order doesn't matter, they're located by name:
    Product name, Store name, Lowest quoted 1 pcs price,
    Highest quoted 1 pcs price, Quotation approval date,
    Supplier name, Supplier name (Accepted), SKU, Quoted variants

--store should be the myshopify.com subdomain of the target shop (e.g. for
bycheri.myshopify.com pass "bycheri"). Get this by running, in the Shopify
MCP: graphql_query with { shop { myshopifyDomain } } -- the "Store name"
column in these exports is that subdomain, not the storefront's custom
domain (e.g. Zanaro Berlin's storefront is zanaroberlin.com, but its
myshopify domain -- and the value in "Store name" -- is bycheri).

Output JSON (written to --out, default parsed_quotations.json):
{
  "store": "bycheri",
  "row_count": 4526,
  "by_title": { "<exact product name>": [<row dict>, ...], ... },
  "by_brand_key": { "<normalized core name>": [<row dict>, ...], ... }
}

Each row dict has: title, lowest, highest, sku, supplier, supplier_accepted,
approval_date, quoted_variants.

Where a product name repeats (re-quotes, renegotiations), all matching rows
are kept in list order; callers should generally pick the one with the most
recent approval_date when a single answer is needed -- see
latest_row(rows) below, importable from this module.
"""
import argparse
import json
import re
import sys


def install_openpyxl_if_needed():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl"], check=True)


def brand_key(title: str) -> str:
    """Reduce a promo-heavy product title to its distinctive core name so
    titles that differ only in marketing text/language still match.

    These xlsx exports and Shopify's own titles both follow a
    "<promo prefix> | <Brand>(TM) <description>" pattern (prefixes like
    "50% RABATT NUR HEUTE | " or "1+1 GRATIS HEUTE | ", in whatever language
    the campaign is running in that week), so the part worth matching on is
    the brand name right after the last "|" and before the (TM) mark.
    """
    seg = title.split('|')[-1].strip()
    m = re.match(r"^[^™]*™", seg)  # up to and including (TM)
    if m:
        core = m.group(0).replace('™', '').strip()
    else:
        core = seg.split(' - ')[0].strip()
        if re.match(r'^\d', core):
            core = re.split(r'[\d]', core)[0].strip()
    core = re.sub(r'[^A-Za-z0-9&]', '', core)
    return core.lower()


def latest_row(rows):
    """Pick the most-recently-approved quote among duplicate rows for the
    same product (falls back to the last row if approval_date is missing)."""
    return sorted(rows, key=lambda r: (r.get('approval_date') or ''), reverse=True)[0]


def parse(xlsx_path: str, store: str):
    install_openpyxl_if_needed()
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    def col(name, *alts):
        for n in (name, *alts):
            if n in idx:
                return idx[n]
        raise KeyError(f"Column not found in xlsx: {name!r} (also tried {alts!r}). "
                        f"Available columns: {header}")

    c_title = col("Product name")
    c_store = col("Store name")
    c_low = col("Lowest quoted 1 pcs price")
    c_high = col("Highest quoted 1 pcs price")
    c_approval = col("Quotation approval date")
    c_supplier = col("Supplier name")
    c_supplier_acc = col("Supplier name (Accepted)")
    c_sku = col("SKU")
    c_variants = col("Quoted variants")

    by_title = {}
    by_key = {}
    row_count = 0
    stores_seen = {}

    for r in ws.iter_rows(min_row=2, values_only=True):
        title = r[c_title]
        if title is None:
            continue

        store_name = r[c_store]
        stores_seen[store_name] = stores_seen.get(store_name, 0) + 1
        if store and store_name != store:
            continue

        row_count += 1
        row = {
            "title": title.strip(),
            "lowest": r[c_low],
            "highest": r[c_high],
            "sku": r[c_sku],
            "supplier": r[c_supplier_acc] or r[c_supplier],
            "approval_date": str(r[c_approval]) if r[c_approval] else None,
            "quoted_variants": r[c_variants],
        }
        by_title.setdefault(row["title"], []).append(row)
        by_key.setdefault(brand_key(row["title"]), []).append(row)

    return {
        "store": store,
        "row_count": row_count,
        "stores_seen": stores_seen,
        "by_title": by_title,
        "by_brand_key": by_key,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx_path")
    ap.add_argument("--store", required=False, default=None,
                     help="myshopify subdomain to filter to, e.g. 'bycheri'. "
                          "Omit to see all store names present (useful for a first pass).")
    ap.add_argument("--out", default="parsed_quotations.json")
    args = ap.parse_args()

    result = parse(args.xlsx_path, args.store)

    if not args.store:
        print("No --store given. Store names found in this file:")
        for name, count in sorted(result["stores_seen"].items(), key=lambda kv: -kv[1]):
            print(f"  {name!r}: {count} rows")
        print("\nRe-run with --store <name> to filter and produce the parsed JSON.")
        return

    print(f"Store {args.store!r}: {result['row_count']} rows "
          f"({len(result['by_title'])} unique titles, {len(result['by_brand_key'])} unique brand keys)")

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
